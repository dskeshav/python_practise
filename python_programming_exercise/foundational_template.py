import os
import openpyxl
import random


# with open('C:\Users\dkeshav\Desktop\ACTR\FoundationalTemplates\FoundationalTemplates\Foundational_data1.xlsx','rb') as foundation:
# wb=openpyxl.load_workbook('C:\Users\dkeshav\Desktop\ACTR\FoundationalTemplates\FoundationalTemplates\Foundational_data1.xlsx')

# print(os.listdir('C:\\Users\\dkeshav\\Desktop\\ACTR\\FoundationalTemplates\\FoundationalTemplates'))   
writeWB=openpyxl.Workbook()
for loop in range(1,61,1):
    wb=openpyxl.load_workbook('C:\\Users\\dkeshav\\Desktop\\ACTR\\FoundationalTemplates\\FoundationalTemplates\\Foundational_data{}.xlsx'.format(loop))
    sheet=wb.active
    datalist=[]
    for fexceldata in range(2,sheet.max_row+1,1):
        datalist.append(sheet.cell(row=fexceldata,column=1).value)
    
    newsheet=writeWB.active
    for iterate in range(1,40001,1):
        newsheet.cell(row=iterate,column=loop).value=random.choice(datalist)
        
        
# writeWB.save('outputexcel.xlsx')    
# print('New file')


filepath='E:\\keshav\\python_programing\\python_practise\\python_programming_exercise\\foundationaldata.xlsx'

writeWB.save(filepath)