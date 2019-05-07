import os 
import openpyxl
import random


wb=openpyxl.load_workbook('AmericanAirlines(2).xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
wb2=openpyxl.load_workbook('ArtoonHireHeadcount (1).xlsx')
worksheet=wb2.get_active_sheet()

exceldata=[]
for rowNum in range(2,sheet.max_row+1,1):
        exceldata.append(sheet.cell(row=rowNum,column=1).value)
print(exceldata)
for randSelection in range(1,10,1):
    randsel=random.sample(exceldata,1)
    worksheet.write(randSelection+1,1,randsel)
    
wb2.close()